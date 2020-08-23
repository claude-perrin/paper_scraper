import openpyxl
from openpyxl import Workbook


def open_excel_file():
    try:
        wb = openpyxl.load_workbook('Zickzackfalz.xlsx')
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    return wb, ws


def find_last_empty_cell():
    workbook = open_excel_file()
    last_row = workbook[1].max_row
    print(last_row)
    return last_row


def insert_word_into_excel(list_of_information):
    empty_cell = find_last_empty_cell() + 1
    wb_sheet_obj = open_excel_file()
    print('=====')
    print('=====')
    for subcategory_list in list_of_information[1:]:
        for products in subcategory_list:
            for title_and_products in products[1:]:
                for product in title_and_products:
                    for product_options in product[2:]:
                        for deliveries in product_options[1:]:
                            for delivery in deliveries:
                                for prices in delivery[1:]:
                                    for price in prices:
                                        try:
                                            paper_name = product[0]
                                            paper_type_name = product_options[0]
                                            delivery_type = delivery[0]
                                            link = product[1]
                                            wb_sheet_obj[1].cell(empty_cell, 1).value = list_of_information[
                                                0]  # Group_name
                                            wb_sheet_obj[1].cell(empty_cell, 2).value = products[0]  # subcat_name
                                            wb_sheet_obj[1].cell(empty_cell, 3).value = paper_name  # Paper_name
                                            wb_sheet_obj[1].cell(empty_cell,
                                                                 4).value = paper_type_name  # Paper_type_name
                                            wb_sheet_obj[1].cell(empty_cell,
                                                                 5).value = delivery_type  # Delivery type
                                            wb_sheet_obj[1].cell(empty_cell, 6).value = price  # Price
                                            wb_sheet_obj[1].cell(empty_cell, 7).value = link  # Link
                                            empty_cell += 1


                                        except Exception as exc:

                                            print(exc)
                                            print(1)
                                            pass
    wb_sheet_obj[0].save("Broschüren mit Spiral-Wire-O-Bindung.xlsx")
